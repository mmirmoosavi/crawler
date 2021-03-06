from time import sleep, time
from tqdm import tqdm

from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from openpyxl import Workbook
import os
from datetime import datetime, timedelta


class TwitterCrawler():
    TWITTER_URL = 'https://twitter.com/'
    SCROLL_MAX_PAUSE_TIME = 3
    FIRST_LOAD = 10
    PROXY = "127.0.0.1:1081"

    def setUp(self, proxy=False):
        options = ChromeOptions()
        if proxy:
            options.add_argument('--proxy-server=socks://%s' % self.PROXY)
        chrome_prefs = {}
        options.experimental_options["prefs"] = chrome_prefs
        chrome_prefs["profile.default_content_settings"] = {"images": 2}
        chrome_prefs["profile.managed_default_content_settings"] = {"images": 2}

        options.add_argument("--start-maximized")

        self.driver = Chrome(executable_path='../drivers/chromedriver', options=options)

    def search_twitter(self, **kwargs):

        keys = kwargs.keys()
        since_date = None
        end_date = None
        search_query = ''
        if 'words' in keys:
            search_query += '{} '.format(kwargs['words'])
        if 'exact_phrases' in keys:
            search_query += '"{}" '.format(kwargs['exact_phrases'])
        if 'hashtags' in keys:
            search_query += '({}) '.format(kwargs['hashtags'])
        if 'extra':
            search_query += '{} '.format(kwargs['extra'])
        if 'since' in twitter_adv_search_parameters.keys():
            year, month, day = tuple(twitter_adv_search_parameters['since'].split('-'))
            since_date = datetime(year=int(year), month=int(month), day=int(day))
        if 'until' in twitter_adv_search_parameters.keys():
            year, month, day = tuple(twitter_adv_search_parameters['until'].split('-'))
            end_date = datetime(year=int(year), month=int(month), day=int(day))
        extracted_posts = []
        if since_date and end_date:
            date = since_date
            temp_query = search_query
            while date <= end_date and date <= datetime.now():
                try:
                    time_query = 'since:{} until:{}'.format(str(date)[:10], str(date + timedelta(days=1))[:10])
                    search_query += time_query
                    print(search_query)
                    self.driver.get(
                        '{}search?q={}'.format(self.TWITTER_URL, search_query))
                    ## ensure the page is fully loaded
                    date += timedelta(days=1)
                    start_time = time()
                    WebDriverWait(self.driver, self.FIRST_LOAD).until(expected_conditions.presence_of_element_located(
                        (By.XPATH, '//article/div/div/div/div[2]/div[2]/div[2]/div[1]/div')))
                    end_time = time()
                    print(end_time - start_time)
                    extracted_posts.extend(self.extract_all_posts())
                    search_query = temp_query
                except:
                    search_query = temp_query
                    continue
        elif since_date:
            date = since_date
            search_query += 'since: {} '.format(date)
            while date <= datetime.now().date():
                self.driver.get(
                    '{}search?q={}'.format(self.TWITTER_URL, search_query))
                ## ensure the page is fully loaded
                start_time = time()
                WebDriverWait(self.driver, self.FIRST_LOAD).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, '//article/div/div/div/div[2]/div[2]/div[2]/div[1]/div')))
                end_time = time()
                print(end_time - start_time)
                extracted_posts.extend(self.extract_all_posts())
                date += timedelta(days=1)
        else:
            self.driver.get(
                '{}search?q={}'.format(self.TWITTER_URL, search_query))
            ## ensure the page is fully loaded
            start_time = time()
            WebDriverWait(self.driver, self.FIRST_LOAD).until(expected_conditions.presence_of_element_located(
                (By.XPATH, '//article/div/div/div/div[2]/div[2]/div[2]/div[1]/div')))
            end_time = time()
            print(end_time - start_time)
            extracted_posts.extend(self.extract_all_posts())
        return extracted_posts

    def extract_all_posts(self):
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        with tqdm() as pbar:
            data = []
            while True:
                # Scroll down to bottom

                posts = self.driver.find_elements_by_xpath(
                    '//article/div/div/div/div[2]/div[2]/div[2]/div[1]/div')
                post_details = self.driver.find_elements_by_xpath(
                    '//article/div/div/div/div[2]/div[2]/div[1]/div/div/div[1]/a')

                print('post deatils: ', len(post_details))
                print('posts_length : ', len(posts))
                post_ids = []
                if len(posts) == len(post_details):
                    for post, detail in zip(posts, post_details):
                        try:
                            href = detail.get_attribute('href')
                            if 'status' not in href:
                                raise Exception
                            splited_href = href.split('/')
                            post_id = splited_href[-1]
                            if post_id in post_ids:
                                raise Exception
                            post_ids.append(post_id)
                            user_name = splited_href[3]
                            text = post.text
                        except:
                            continue

                        data.append([post_id, user_name, text, href])
                print(len(data))
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

                # Wait to load page
                sleep(self.SCROLL_MAX_PAUSE_TIME)

                # Calculate new scroll height and compare with last scroll height
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
                pbar.update()
        return data

    def write_to_excel(self, array):
        head = ['post_id', 'user_name', 'text', 'link']
        wb = Workbook()
        ws = wb.active
        ws.title = 'twitter_data'
        for index, value in enumerate(head, start=1):
            cell = ws.cell(row=1, column=index)
            cell.value = value
        for row in range(len(array)):
            for col in range(len(head)):
                cell = ws.cell(row=row + 2, column=col + 1)
                cell.value = array[row][col]

        os.makedirs('../data', exist_ok=True)
        wb.save('../data/twitter_data.xlsx')

    def load_page(self):
        while True:
            is_ready = self.driver.execute_script('return document.readyState;')
            sleep(0.1)
            print(is_ready)
            if is_ready == 'complete':
                break
        return

    def tearDown(self):
        self.driver.close()


if __name__ == '__main__':
    crawler = TwitterCrawler()
    crawler.setUp(True)
    twitter_adv_search_parameters = {'words': 'چکاپا', 'since': '2020-07-22',
                                     'until': '2020-11-10', 'extra': 'lang:fa -filter:replies'}

    data = crawler.search_twitter(**twitter_adv_search_parameters)
    crawler.tearDown()
    crawler.write_to_excel(array=data)
